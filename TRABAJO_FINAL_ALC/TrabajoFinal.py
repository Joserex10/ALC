import tkinter as tk                                                # Para crear interfaces graficas en Python
from tkinter import filedialog, messagebox                          # filedialog para seleccionar archivos, messagebox para mostrar mensajes al usuario
import numpy as np                                                  # Para trabajar con arreglos y operaciones matematicas
import pandas as pd                                                 # Para manipulacion y analisis de datos
from sklearn.decomposition import PCA                               # Modulo para aplicar Analisis de Componentes Principales (PCA)
import matplotlib.pyplot as plt                                     # Para crear graficos
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg     # Integracion de graficos de Matplotlib en Tkinter
from mpl_toolkits.mplot3d import Axes3D                             # Modulo para graficos 3D en Matplotlib
from tkinter import simpledialog                                    # Modulo para dialogos de entrada de texto en Tkinter
import openpyxl                                                     # Para manipulacion de archivos de Excel
from openpyxl.chart import BarChart, Reference                      # Clases para crear graficos de barras en Excel
from openpyxl.styles import PatternFill                             # Clase para aplicar estilos de relleno en celdas de Excel

# Variables globales para almacenar datos de PCA
data, principal_components, autovalores, autovectores, pca = None, None, None, None, None

# Convierte matrices de valores en una cadena de texto formateada
def format_values(values, is_vector=False):
    if is_vector:
        formatted = '\n'.join(['  '.join([f'{num:.4f}' for num in row]) for row in values])
    else:
        formatted = ', '.join([f'{num:.4f}' for num in values])
    return formatted


# Funcion para cargar y aplicar PCA a un archivo CSV
def load_and_apply_pca():
    global data, principal_components, autovalores, autovectores, pca
    file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
    if not file_path:
        return
        
    try:
        data = pd.read_csv(file_path, header=None)

        if not np.all(np.isreal(data.values)):
            messagebox.showerror("Error", "El archivo debe contener únicamente datos numéricos para aplicar PCA.")
            return

        if data.shape[1] < 3:
            messagebox.showerror("Error", "El archivo debe tener al menos 3 columnas para graficar en 3D.")
            return

        n_components = select_components(data.shape[1])
        if n_components is None:
            return

        pca = PCA(n_components=n_components)
        principal_components = pca.fit_transform(data)
        autovalores = pca.explained_variance_
        autovectores = pca.components_
        
        create_result_window()
        
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrio un error al aplicar PCA:\n{e}")

# Funcion para mostrar los datos originales en 3D
def show_original_data_3d():
    original_window = tk.Toplevel(result_window)
    original_window.title("Datos Originales - Grafico 3D")
    
    fig = plt.figure(figsize=(8, 6))
    ax = fig.add_subplot(111, projection='3d')
    
    ax.scatter(data.iloc[:, 0], data.iloc[:, 1], data.iloc[:, 2], c='blue', marker='o', s=50)
    ax.set_title("Datos Originales en 3D")
    ax.set_xlabel("Variable 1")
    ax.set_ylabel("Variable 2")
    ax.set_zlabel("Variable 3")

    canvas = FigureCanvasTkAgg(fig, master=original_window)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10, fill=tk.BOTH, expand=True)

# Funcion para mostrar autovalores y autovectores
def show_eigenvalues_vectors():
    autovalores_str = "Autovalores:\n" + format_values(autovalores)
    autovectores_str = "Autovectores:\n" + format_values(autovectores, is_vector=True)
    
    text_window = tk.Toplevel(result_window)
    text_window.title("Autovalores y Autovectores")
    text_box = tk.Text(text_window, wrap=tk.WORD, width=80, height=20)
    text_box.insert(tk.END, autovalores_str + "\n\n" + autovectores_str)
    text_box.config(state=tk.DISABLED)
    text_box.pack(pady=10)

# Funcion para mostrar el grafico de varianza explicada
def plot_explained_variance():
    explained_variance = pca.explained_variance_ratio_ * 100
    components = np.arange(1, len(explained_variance) + 1)
    
    fig, ax = plt.subplots(figsize=(8, 6))
    bars = ax.bar(components, explained_variance, color='cyan')
    ax.set_title('Varianza Explicada por Cada Componente')
    ax.set_xlabel('Componente Principal')
    ax.set_ylabel('Varianza Explicada (%)')
    ax.set_xticks(components)
    
    for bar in bars:
        height = bar.get_height()
        ax.annotate(f'{height:.2f}%',  
                    xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3),  
                    textcoords="offset points",
                    ha='center', va='bottom')

    var_window = tk.Toplevel(result_window)
    var_window.title("Varianza Explicada")
    canvas = FigureCanvasTkAgg(fig, master=var_window)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10, fill=tk.BOTH, expand=True)

# Funcion para mostrar los datos transformados en un grafico de dispersion
def plot_principal_components():
    fig, ax = plt.subplots(figsize=(8, 6))
    scatter = ax.scatter(principal_components[:, 0], principal_components[:, 1], 
                         c=['red' if x > 0 else 'blue' for x in principal_components[:, 0]], edgecolor='k', s=50)
    ax.set_title("Datos proyectados en los dos primeros componentes principales")
    ax.set_xlabel("Componente Principal 1 (PC1)")
    ax.set_ylabel("Componente Principal 2 (PC2)")
    ax.grid(True)

    pc_window = tk.Toplevel(result_window)
    pc_window.title("Componentes Principales")
    canvas = FigureCanvasTkAgg(fig, master=pc_window)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10, fill=tk.BOTH, expand=True)

# Funcion para guardar los datos transformados
def save_transformed_data():
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if save_path:
        num_components = principal_components.shape[1]
        
        # Crear DataFrame con los componentes principales y autovalores
        transformed_df = pd.DataFrame(np.round(principal_components, 3), columns=[f'PC{i+1}' for i in range(num_components)])
        transformed_df[f'Autovalores (PC1-PC{num_components})'] = ''
        for i, val in enumerate(autovalores):
            transformed_df.loc[i, f'Autovalores (PC1-PC{num_components})'] = f"PC{i+1}: {val:.4f}"

        # Crear DataFrame para la varianza explicada
        explained_variance_df = pd.DataFrame({
            'Componente': np.arange(1, len(pca.explained_variance_ratio_) + 1),
            'Varianza Explicada': pca.explained_variance_ratio_
        })

        # Guardar el archivo Excel con dos hojas
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            transformed_df.to_excel(writer, index=False, sheet_name='Datos Transformados')
            explained_variance_df.to_excel(writer, index=False, sheet_name='Varianza Explicada')

        # Aplicar colores a las celdas en "Datos Transformados"
        wb = openpyxl.load_workbook(save_path)
        sheet = wb["Datos Transformados"]
        fill_color = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for row in sheet['A2:B' + str(sheet.max_row)]:  # Aplicar color solo a las celdas de PC1 y PC2
            for cell in row:
                cell.fill = fill_color

        # Crear grafico de barras en "Varianza Explicada"
        sheet_varianza = wb["Varianza Explicada"]
        chart = BarChart()
        data = Reference(sheet_varianza, min_col=2, min_row=1, max_row=sheet_varianza.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.title = "Varianza Explicada por Componente"
        chart.x_axis.title = "Componente Principal"
        chart.y_axis.title = "Varianza Explicada (%)"
        sheet_varianza.add_chart(chart, "E5")

        # Guardar el libro de Excel con todas las modificaciones
        wb.save(save_path)
        
        messagebox.showinfo("Guardado", f"Datos transformados y graficos guardados en: {save_path}")


# Funcion para seleccionar el número de componentes
def select_components(max_components):
    try:
        value = int(simpledialog.askstring("Número de Componentes", f"Ingresa el número de componentes principales (1-{max_components}):"))
        if 1 <= value <= max_components:
            return value
        else:
            messagebox.showerror("Error", f"El número de componentes debe estar entre 1 y {max_components}.")
            return None
    except ValueError:
        messagebox.showerror("Error", "Debes ingresar un número valido.")
        return None

# Funcion para crear la ventana de resultados con cinco botones
def create_result_window():
    global result_window
    result_window = tk.Toplevel(root)
    result_window.title("Resultados de PCA")
    result_window.geometry("400x400")
    
    tk.Button(result_window, text="Ver Datos Originales en 3D", command=show_original_data_3d).pack(pady=10)
    tk.Button(result_window, text="Ver Autovalores y Autovectores", command=show_eigenvalues_vectors).pack(pady=10)
    tk.Button(result_window, text="Ver Varianza Explicada", command=plot_explained_variance).pack(pady=10)
    tk.Button(result_window, text="Ver Componentes Principales", command=plot_principal_components).pack(pady=10)
    tk.Button(result_window, text="Guardar Datos Transformados", command=save_transformed_data).pack(pady=10)

# Crear la ventana principal
root = tk.Tk()
root.title("Analizador de Archivos - PCA")
root.geometry("600x600")

bg_color = "#1e1e1e"  
btn_color = "#5bc0de"
text_color = "#f8f9fa"

root.configure(bg=bg_color)

title_label = tk.Label(root, text="Analizador de Archivos - PCA", font=("Helvetica", 18, "bold"), bg=bg_color, fg=text_color)
title_label.pack(pady=50)

pca_btn = tk.Button(root, text="Cargar Archivo CSV y Aplicar PCA", font=("Arial", 14), bg=btn_color, fg=text_color, padx=10, pady=5)
pca_btn.config(bd=0, relief=tk.FLAT)
pca_btn.config(command=load_and_apply_pca)
pca_btn.pack(padx=10, pady=50)

root.mainloop()
